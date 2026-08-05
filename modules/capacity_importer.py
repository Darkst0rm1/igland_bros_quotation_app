"""Import bundles-per-container from a logistics workbook.

Same shape of problem as the price-list importer, so the same approach: detect
the header row rather than assume it (the reference file has a title bar on row
1 and headers on row 3), normalise the headings, and report what it found
before writing anything.

**Anomalies are flagged, not corrected.** Capacity should fall as boxes get
bigger. In the reference file it does — 3,500 bundles for a 7" down to 890 for
an 18" — and then rises to 1,512 for the 20", which is roughly 3.5x the trend
where every other size sits between 1.0x and 1.7x. That is either a different
bundle size for that item or a transcription error, and it is not this
module's place to guess which. The row imports with ``is_anomalous`` set and a
note explaining the discrepancy.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from decimal import Decimal
from typing import BinaryIO

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from modules.audit_service import record_audit
from modules.authorization import AuthUser, require
from modules.calculation_engine import to_decimal
from modules.constants import (
    DEFAULT_CONTAINER_SIZE,
    DEFAULT_CONTAINER_TYPE,
    AuditAction,
    ContainerSize,
    ContainerType,
    EntityType,
    Perm,
)
from modules.models import ProductContainerCapacity
from modules.validation import clean_text, coerce_decimal

log = logging.getLogger(__name__)


class CapacityImportError(ValueError):
    """The workbook could not be read. Safe to show the user."""


#: Heading synonyms. Only two columns matter, but the file is hand-maintained
#: so the wording will drift.
_PRODUCT_HEADERS = {"product", "size", "item", "description"}
_BUNDLE_HEADERS = {
    "bundles per container", "bundles/container", "bundles per ctnr",
    "bundles", "bundles per 40hc", "bundles per container 40hc",
}
_UNITS_HEADERS = {
    "units per bundle", "boxes per bundle", "pieces per bundle", "units/bundle",
}
_PALLET_HEADERS = {"pallets per container", "pallets", "pallets/container"}

#: A row is only a header if it names both the things we need.
_MIN_MATCHES = 2


@dataclass
class CapacityRow:
    source_row_no: int
    product_label: str
    bundles_per_container: Decimal
    units_per_bundle: Decimal | None = None
    pallets_per_container: Decimal | None = None
    error: str | None = None
    is_anomalous: bool = False
    anomaly_note: str | None = None

    @property
    def ok(self) -> bool:
        return self.error is None


@dataclass
class CapacityPlan:
    header_row: int
    rows: list[CapacityRow]
    container_size: ContainerSize
    container_type: ContainerType
    notes: list[str] = field(default_factory=list)

    def counts(self) -> dict[str, int]:
        return {
            "ok": sum(1 for r in self.rows if r.ok),
            "error": sum(1 for r in self.rows if not r.ok),
            "anomalous": sum(1 for r in self.rows if r.is_anomalous),
        }


def _normalise(value: object) -> str | None:
    text = clean_text(value)
    return text.casefold() if text else None


def _size_from_note(text: str) -> ContainerSize:
    """Read the container the figures apply to out of the sheet's own note.

    The reference file says "Container type: 40' HC", and the numbers are
    meaningless without knowing which container they describe — so this is read
    rather than assumed, and falls back to the configured default.
    """
    lowered = text.casefold()
    if "45" in lowered and ("hc" in lowered or "high cube" in lowered):
        return ContainerSize.FORTY_FIVE_FT_HC
    if "40" in lowered and ("hc" in lowered or "high cube" in lowered):
        return ContainerSize.FORTY_FT_HC
    if "40" in lowered:
        return ContainerSize.FORTY_FT
    if "20" in lowered:
        return ContainerSize.TWENTY_FT
    return DEFAULT_CONTAINER_SIZE


def read_workbook(
    source: str | BinaryIO, sheet_name: str | None = None
) -> CapacityPlan:
    """Parse the workbook. Reads only; writes nothing."""
    workbook = load_workbook(source, data_only=True)
    try:
        if sheet_name and sheet_name not in workbook.sheetnames:
            raise CapacityImportError(f"The workbook has no sheet named {sheet_name!r}.")
        sheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]

        max_row = sheet.max_row or 0
        max_col = sheet.max_column or 0

        header_row = None
        columns: dict[str, int] = {}
        for row in range(1, max_row + 1):
            found: dict[str, int] = {}
            for col in range(1, max_col + 1):
                heading = _normalise(sheet.cell(row=row, column=col).value)
                if heading is None:
                    continue
                if heading in _PRODUCT_HEADERS:
                    found["product"] = col
                elif heading in _BUNDLE_HEADERS:
                    found["bundles"] = col
                elif heading in _UNITS_HEADERS:
                    found["units"] = col
                elif heading in _PALLET_HEADERS:
                    found["pallets"] = col
            if len(found) >= _MIN_MATCHES and "product" in found and "bundles" in found:
                header_row, columns = row, found
                break

        if header_row is None:
            raise CapacityImportError(
                "No capacity table was found. A header row must name a product "
                "column and a bundles-per-container column."
            )

        # Free text below the table tells us which container the figures are for.
        trailing = " ".join(
            clean_text(sheet.cell(row=r, column=1).value) or ""
            for r in range(header_row, max_row + 1)
        )
        container_size = _size_from_note(trailing)

        rows: list[CapacityRow] = []
        for row_no in range(header_row + 1, max_row + 1):
            label = clean_text(sheet.cell(row=row_no, column=columns["product"]).value)
            raw_bundles = sheet.cell(row=row_no, column=columns["bundles"]).value
            if not label:
                continue
            if raw_bundles is None:
                # A trailing note occupies the product column; stop at it.
                continue
            try:
                bundles = coerce_decimal(raw_bundles)
                if bundles is None or bundles <= 0:
                    raise ValueError("bundles per container must be greater than zero")
                units = (
                    coerce_decimal(sheet.cell(row=row_no, column=columns["units"]).value)
                    if "units" in columns else None
                )
                pallets = (
                    coerce_decimal(sheet.cell(row=row_no, column=columns["pallets"]).value)
                    if "pallets" in columns else None
                )
                rows.append(
                    CapacityRow(row_no, label, bundles, units, pallets)
                )
            except (ValueError, ArithmeticError) as exc:
                rows.append(
                    CapacityRow(row_no, label, Decimal("0"), error=str(exc))
                )

        _flag_anomalies(rows)

        notes = [
            clean_text(sheet.cell(row=r, column=1).value)
            for r in range(header_row + len(rows) + 1, max_row + 1)
        ]
        return CapacityPlan(
            header_row=header_row,
            rows=rows,
            container_size=container_size,
            container_type=DEFAULT_CONTAINER_TYPE,
            notes=[n for n in notes if n],
        )
    finally:
        workbook.close()


def _size_in_inches(label: str) -> Decimal | None:
    match = re.search(r"(\d+(?:\.\d+)?)", label or "")
    return Decimal(match.group(1)) if match else None


def _flag_anomalies(rows: list[CapacityRow]) -> None:
    """Mark rows whose capacity contradicts the trend of the others.

    Capacity should fall as the footprint grows. A row that carries materially
    more than its neighbours usually means its bundle holds a different number
    of units — worth a human deciding, so it is flagged rather than adjusted.
    """
    usable = [
        (r, _size_in_inches(r.product_label))
        for r in rows if r.ok
    ]
    ordered = [(r, s) for r, s in usable if s is not None]
    if len(ordered) < 3:
        return
    ordered.sort(key=lambda pair: pair[1])

    for index, (row, size) in enumerate(ordered):
        if index == 0:
            continue
        previous_row, previous_size = ordered[index - 1]
        if previous_row.bundles_per_container <= 0 or size <= previous_size:
            continue
        # A bigger box holding more per container is the contradiction.
        if row.bundles_per_container > previous_row.bundles_per_container:
            expected = (
                previous_row.bundles_per_container
                * (previous_size * previous_size)
                / (size * size)
            )
            row.is_anomalous = True
            row.anomaly_note = (
                f"{row.bundles_per_container:,.0f} bundles is more than the "
                f"{previous_row.bundles_per_container:,.0f} recorded for the smaller "
                f'{previous_size:g}" size. Scaling by footprint would suggest roughly '
                f"{expected:,.0f}. Imported as given — check whether this bundle holds "
                f"a different number of units."
            )


# --------------------------------------------------------------------------- #
# Commit
# --------------------------------------------------------------------------- #

def commit(
    session: Session,
    user: AuthUser,
    plan: CapacityPlan,
    file_name: str,
    *,
    container_size: ContainerSize | None = None,
    container_type: ContainerType | None = None,
    units_per_bundle: Decimal | None = None,
) -> dict[str, int]:
    """Write the capacities, matching rows to products by size label.

    ``units_per_bundle`` may be supplied for the whole workbook when the source
    does not state it. Left as ``None``, pieces and cases per container stay
    unavailable rather than being derived from a guess.
    """
    require(user, Perm.PRICE_IMPORT)

    from modules.repositories import find_product_by_size

    size = container_size or plan.container_size
    ctype = container_type or plan.container_type

    created = updated = skipped = failed = 0
    for row in plan.rows:
        if not row.ok:
            failed += 1
            continue

        product = find_product_by_size(session, row.product_label)
        if product is None:
            skipped += 1
            log.warning(
                "No product matches %r; capacity row %d skipped",
                row.product_label, row.source_row_no,
            )
            continue

        existing = session.execute(
            _capacity_query(product.id, size, ctype)
        ).scalar_one_or_none()

        target = existing or ProductContainerCapacity(
            product_id=product.id, container_size=size, container_type=ctype
        )
        target.bundles_per_container = row.bundles_per_container
        if row.units_per_bundle is not None:
            target.units_per_bundle = row.units_per_bundle
        elif units_per_bundle is not None:
            target.units_per_bundle = to_decimal(units_per_bundle)
        if row.pallets_per_container is not None:
            target.pallets_per_container = row.pallets_per_container
        target.source_workbook_name = file_name
        target.source_row_no = row.source_row_no
        target.is_anomalous = row.is_anomalous
        target.anomaly_note = row.anomaly_note
        target.created_by_id = user.id

        if existing is None:
            session.add(target)
            created += 1
        else:
            updated += 1

    session.flush()
    summary = {
        "created": created, "updated": updated,
        "skipped": skipped, "failed": failed,
        "anomalous": plan.counts()["anomalous"],
    }
    record_audit(
        session, user, AuditAction.CONTAINER_CAPACITY_IMPORTED,
        EntityType.PRODUCT_CONTAINER_CAPACITY, None,
        new_value={
            "file": file_name,
            "container": f"{size.value}/{ctype.value}",
            **summary,
        },
    )
    log.info("Capacity import: %s", summary)
    return summary


def _capacity_query(product_id: int, size: ContainerSize, ctype: ContainerType):  # noqa: ANN201
    from sqlalchemy import select

    return select(ProductContainerCapacity).where(
        ProductContainerCapacity.product_id == product_id,
        ProductContainerCapacity.container_size == size,
        ProductContainerCapacity.container_type == ctype,
    )
