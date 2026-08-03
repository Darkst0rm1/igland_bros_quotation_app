"""Price-list workbook importer.

Pipeline: **detect → normalise → validate → diff → preview → commit**.

The first four stages are pure: they read a workbook and produce a plan without
touching the database, which is what makes the preview honest — what the
operator approves is exactly what gets written.

Three properties of the reference workbook drive the design, all confirmed by
inspection (docs/PHASE1_REFERENCE_ANALYSIS.md):

1. **The header row is not row 1, and there is more than one.** The reference
   file has headers at rows 2 and 26. Header rows are therefore detected, not
   assumed.
2. **Board quality is a per-row value.** The "alternative quality" section
   contains two different qualities (rows 27–33 and 34–38). The section heading
   is recorded for the audit summary and never used to infer quality.
3. **Headers contain embedded newlines** (``"Standard\\nPrice/Pack"``), and the
   tier columns follow a pattern, so they are matched by regex rather than by a
   fixed lookup — a future twelve-container tier needs no code change.
"""

from __future__ import annotations

import datetime as dt
import logging
import re
from dataclasses import dataclass, field
from decimal import Decimal
from typing import Any, BinaryIO

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError
from sqlalchemy.orm import Session

from modules.audit_service import record_audit
from modules.constants import (
    AuditAction,
    EntityType,
    ImportJobStatus,
    ImportRowAction,
    ImportRowStatus,
)
from modules.models import ImportJob, ImportRow, Product, ProductPrice, ProductVariant
from modules.repositories import (
    find_variant_by_natural_key,
    price_tier_map,
    supersede_price,
)
from modules.validation import PriceRowInput, clean_text, coerce_decimal

log = logging.getLogger(__name__)


class ImportError_(Exception):
    """Raised for a workbook that cannot be read or has no usable table."""


# --------------------------------------------------------------------------- #
# Header normalisation
# --------------------------------------------------------------------------- #

#: Identity columns.
_IDENTITY_HEADERS: dict[str, str] = {
    "product": "product",
    "item": "product",
    "size": "product",
    "depth": "depth",
    "flute": "flute",
    "case": "case_pack",
    "case pack": "case_pack",
    "casepack": "case_pack",
    "quality": "board_quality",
    "board quality": "board_quality",
}

#: Price columns. ``standard`` or ``<n> container(s)``, then pack or piece.
_PRICE_HEADER = re.compile(
    r"^(?:(?P<standard>standard)|(?P<n>\d+)\s*containers?)\s*"
    r"price\s*/\s*(?P<basis>pack|pcs|piece|pieces)$"
)

#: Spelled-out prefixes for the tier field names.
_NUMBER_WORDS: dict[int, str] = {
    1: "one", 2: "two", 3: "three", 4: "four", 5: "five", 6: "six",
    7: "seven", 8: "eight", 9: "nine", 10: "ten", 11: "eleven", 12: "twelve",
}

#: ``<n> containers`` → price-tier code, for the tiers the catalogue supports.
CONTAINER_TIER_CODES: dict[int, str] = {
    3: "THREE_CONTAINER",
    8: "EIGHT_CONTAINER",
}


def normalise_header(raw: Any) -> str | None:
    """Map one header cell to a field name, or ``None`` if unrecognised.

    ``"Standard\\nPrice/Pack"`` → ``standard_price_per_pack``
    ``"3 containers\\nPrice/Pcs"`` → ``three_container_price_per_piece``
    """
    text = clean_text(raw)
    if not text:
        return None
    key = text.casefold()

    if key in _IDENTITY_HEADERS:
        return _IDENTITY_HEADERS[key]

    match = _PRICE_HEADER.match(key)
    if not match:
        return None

    basis = "pack" if match.group("basis") == "pack" else "piece"
    if match.group("standard"):
        return f"standard_price_per_{basis}"

    count = int(match.group("n"))
    word = _NUMBER_WORDS.get(count)
    if word is None:  # a tier this catalogue has no name for
        return None
    return f"{word}_container_price_per_{basis}"


def _is_price_field(name: str) -> bool:
    return name.endswith(("_price_per_pack", "_price_per_piece"))


# --------------------------------------------------------------------------- #
# Detection
# --------------------------------------------------------------------------- #

#: A row must map this many headers to count as a header row. Four is enough to
#: exclude a title bar or a stray label while still matching a table that omits
#: optional columns.
_MIN_HEADER_MATCHES = 4


@dataclass
class Block:
    """One header row plus the contiguous data rows beneath it."""

    header_row: int
    first_data_row: int
    last_data_row: int
    #: ``{column index: field name}``
    columns: dict[int, str] = field(default_factory=dict)
    section_label: str | None = None

    @property
    def row_count(self) -> int:
        return max(0, self.last_data_row - self.first_data_row + 1)


def list_sheets(source: str | BinaryIO) -> list[str]:
    workbook = load_workbook(source, data_only=True, read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _row_values(sheet: Worksheet, row: int, max_col: int) -> list[Any]:
    return [sheet.cell(row=row, column=col).value for col in range(1, max_col + 1)]


def _detect_header_columns(values: list[Any]) -> dict[int, str]:
    """Map column index → field name for a candidate header row."""
    columns: dict[int, str] = {}
    for index, value in enumerate(values, start=1):
        field_name = normalise_header(value)
        if field_name and field_name not in columns.values():
            columns[index] = field_name
    return columns


def _find_section_label(sheet: Worksheet, before_row: int, after_row: int) -> str | None:
    """The nearest single-cell text row above a header, e.g. "alternative quality".

    Recorded on every imported row for the audit summary. It is deliberately
    never consulted when determining board quality.
    """
    for row in range(before_row - 1, max(after_row, 0), -1):
        first = clean_text(sheet.cell(row=row, column=1).value)
        if not first:
            continue
        # A title bar repeats across the sheet; a section label is distinctive.
        rest = [
            clean_text(sheet.cell(row=row, column=col).value)
            for col in range(2, min(sheet.max_column, 6) + 1)
        ]
        if not any(rest) and normalise_header(first) is None:
            return first
    return None


def detect_blocks(sheet: Worksheet) -> list[Block]:
    """Find every header row and the data rows belonging to it.

    A block ends at the first row with no value in the product column, which is
    what separates the price table from the terms footer beneath it.
    """
    max_row = sheet.max_row or 0
    max_col = sheet.max_column or 0
    blocks: list[Block] = []
    row = 1

    while row <= max_row:
        columns = _detect_header_columns(_row_values(sheet, row, max_col))
        if len(columns) < _MIN_HEADER_MATCHES or "product" not in columns.values():
            row += 1
            continue

        product_col = next(c for c, name in columns.items() if name == "product")
        first_data = row + 1
        last_data = row
        probe = first_data
        while probe <= max_row:
            if clean_text(sheet.cell(row=probe, column=product_col).value) is None:
                break
            last_data = probe
            probe += 1

        if last_data >= first_data:
            previous_end = blocks[-1].last_data_row if blocks else 0
            blocks.append(
                Block(
                    header_row=row,
                    first_data_row=first_data,
                    last_data_row=last_data,
                    columns=columns,
                    section_label=_find_section_label(sheet, row, previous_end),
                )
            )
            row = last_data + 1
        else:
            row += 1

    return blocks


def extract_terms(sheet: Worksheet, blocks: list[Block]) -> dict[str, str]:
    """Label/value pairs from the footer rows beneath each block.

    Returned for display as *suggestions* only. They are never applied to term
    templates automatically — the reference workbook's "Valid through July '26"
    is a stale date that must not become a default.
    """
    terms: dict[str, str] = {}
    max_row = sheet.max_row or 0
    boundaries = [(b.last_data_row, b) for b in blocks]

    for index, (end_row, _block) in enumerate(boundaries):
        # Stop *before* the next block's header row — otherwise "Product | Depth"
        # is read as a term whose label is "Product".
        stop_exclusive = (
            boundaries[index + 1][1].header_row
            if index + 1 < len(boundaries)
            else max_row + 1
        )
        last_label: str | None = None
        for row in range(end_row + 1, stop_exclusive):
            label = clean_text(sheet.cell(row=row, column=1).value)
            value = clean_text(sheet.cell(row=row, column=2).value)
            if value is None:
                continue
            key = label or last_label
            if key is None:
                continue
            # A continuation row repeats its predecessor's label.
            terms[key] = f"{terms[key]} {value}" if key in terms and not label else value
            last_label = key
    return terms


# --------------------------------------------------------------------------- #
# Parsing
# --------------------------------------------------------------------------- #

@dataclass
class ParsedRow:
    source_row_no: int
    section_label: str | None
    raw: dict[str, Any]
    parsed: PriceRowInput | None = None
    error: str | None = None

    @property
    def ok(self) -> bool:
        return self.parsed is not None


def parse_block(sheet: Worksheet, block: Block) -> list[ParsedRow]:
    rows: list[ParsedRow] = []
    for row_no in range(block.first_data_row, block.last_data_row + 1):
        raw = {
            name: sheet.cell(row=row_no, column=col).value
            for col, name in block.columns.items()
        }
        payload: dict[str, Any] = {
            "source_row_no": row_no,
            "section_label": block.section_label,
            **{k: v for k, v in raw.items() if not _is_price_field(k)},
            **{k: v for k, v in raw.items() if _is_price_field(k)},
        }
        try:
            parsed = PriceRowInput(**payload)
            rows.append(ParsedRow(row_no, block.section_label, raw, parsed=parsed))
        except ValidationError as exc:
            message = "; ".join(
                f"{'.'.join(str(p) for p in e['loc'])}: {e['msg']}" for e in exc.errors()
            )
            rows.append(ParsedRow(row_no, block.section_label, raw, error=message))
    return rows


def read_workbook(
    source: str | BinaryIO, sheet_name: str | None = None
) -> tuple[list[Block], list[ParsedRow], dict[str, str]]:
    """Detect blocks, parse every row, and extract the terms footer.

    Pure: reads the file and returns a plan. Nothing is written.
    """
    workbook = load_workbook(source, data_only=True)
    try:
        if sheet_name and sheet_name not in workbook.sheetnames:
            raise ImportError_(f"The workbook has no sheet named {sheet_name!r}.")
        sheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]

        blocks = detect_blocks(sheet)
        if not blocks:
            raise ImportError_(
                "No price table was found on this sheet. A header row must "
                "include at least a Product column plus three of Depth, Flute, "
                "Case, Quality or a price column."
            )

        rows: list[ParsedRow] = []
        for block in blocks:
            rows.extend(parse_block(sheet, block))
        return blocks, rows, extract_terms(sheet, blocks)
    finally:
        workbook.close()


# --------------------------------------------------------------------------- #
# Diff
# --------------------------------------------------------------------------- #

@dataclass
class RowPlan:
    """What the importer proposes to do with one row."""

    row: ParsedRow
    action: ImportRowAction
    status: ImportRowStatus
    variant_id: int | None = None
    product_id: int | None = None
    message: str | None = None
    #: ``{tier_code: (old_price_per_pack | None, new_price_per_pack)}``
    price_changes: dict[str, tuple[Decimal | None, Decimal]] = field(default_factory=dict)

    @property
    def is_error(self) -> bool:
        return self.status is not ImportRowStatus.OK


@dataclass
class ImportPlan:
    blocks: list[Block]
    plans: list[RowPlan]
    terms: dict[str, str]
    effective_from: dt.date

    def counts(self) -> dict[str, int]:
        counts = {"create": 0, "update": 0, "skip": 0, "error": 0, "duplicate": 0}
        for plan in self.plans:
            if plan.status is ImportRowStatus.ERROR:
                counts["error"] += 1
            elif plan.status is ImportRowStatus.DUPLICATE:
                counts["duplicate"] += 1
            elif plan.action is ImportRowAction.CREATE:
                counts["create"] += 1
            elif plan.action is ImportRowAction.UPDATE:
                counts["update"] += 1
            else:
                counts["skip"] += 1
        return counts


def build_plan(
    session: Session,
    rows: list[ParsedRow],
    blocks: list[Block],
    terms: dict[str, str],
    effective_from: dt.date,
    currency: str = "USD",
) -> ImportPlan:
    """Decide create / update / skip for every row, without writing anything."""
    plans: list[RowPlan] = []
    seen: dict[tuple, int] = {}

    for row in rows:
        if not row.ok:
            plans.append(
                RowPlan(row, ImportRowAction.SKIP, ImportRowStatus.ERROR, message=row.error)
            )
            continue

        parsed = row.parsed
        key = parsed.natural_key
        if key in seen:
            plans.append(
                RowPlan(
                    row, ImportRowAction.SKIP, ImportRowStatus.DUPLICATE,
                    message=(
                        f"The same product and board quality already appears on row "
                        f"{seen[key]} of this workbook."
                    ),
                )
            )
            continue
        seen[key] = row.source_row_no

        variant = find_variant_by_natural_key(
            session,
            size_label=parsed.product,
            depth=parsed.depth,
            flute=parsed.flute,
            case_pack=parsed.case_pack,
            board_quality=parsed.board_quality,
        )

        changes: dict[str, tuple[Decimal | None, Decimal]] = {}
        if variant is None:
            for tier_code, (pack, _piece) in parsed.tier_prices().items():
                if pack is not None:
                    changes[tier_code] = (None, pack)
            plans.append(
                RowPlan(row, ImportRowAction.CREATE, ImportRowStatus.OK, price_changes=changes)
            )
            continue

        from modules.repositories import get_effective_price, get_latest_price

        for tier_code, (pack, piece) in parsed.tier_prices().items():
            if pack is None:
                continue
            existing = get_effective_price(
                session, variant.id, tier_code, effective_from, currency
            )
            # Both columns are compared. They are independent values on this
            # catalogue, so a piece price can move while the pack price holds.
            unchanged = (
                existing is not None
                and existing.price_per_pack == pack
                and (piece is None or existing.price_per_piece == piece)
            )
            if not unchanged:
                changes[tier_code] = (
                    existing.price_per_pack if existing else None,
                    pack,
                )

        if not changes:
            plans.append(
                RowPlan(
                    row, ImportRowAction.SKIP, ImportRowStatus.OK,
                    variant_id=variant.id, product_id=variant.product_id,
                    message="Prices are unchanged.",
                )
            )
            continue

        # Only once we know a price genuinely moved does backdating matter.
        # Checking earlier would flag an identical re-import at the same
        # effective date, which is a no-op and must simply skip.
        conflict = next(
            (
                latest
                for tier_code in changes
                if (latest := get_latest_price(session, variant.id, tier_code, currency))
                is not None
                and latest.effective_from >= effective_from
            ),
            None,
        )
        if conflict is not None:
            plans.append(
                RowPlan(
                    row, ImportRowAction.SKIP, ImportRowStatus.ERROR,
                    variant_id=variant.id, product_id=variant.product_id,
                    message=(
                        f"A price for this product is already effective from "
                        f"{conflict.effective_from}. Choose an effective date after "
                        f"that — an issued price cannot be rewritten."
                    ),
                )
            )
            continue

        plans.append(
            RowPlan(
                row, ImportRowAction.UPDATE, ImportRowStatus.OK,
                variant_id=variant.id, product_id=variant.product_id,
                price_changes=changes,
            )
        )

    return ImportPlan(blocks=blocks, plans=plans, terms=terms, effective_from=effective_from)


# --------------------------------------------------------------------------- #
# Commit
# --------------------------------------------------------------------------- #

def _slugify(text: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "-", text).strip("-").upper()


def _item_number_for(parsed: PriceRowInput) -> str:
    return _slugify(parsed.product)[:40] or "ITEM"


def _variant_item_number(parsed: PriceRowInput, product_item_number: str) -> str:
    quality = _slugify(parsed.board_quality)[:30]
    return f"{product_item_number}-{quality}-{parsed.case_pack}"[:80]


def commit_plan(
    session: Session,
    plan: ImportPlan,
    user: Any,
    file_name: str,
    sheet_name: str | None = None,
    storage_key: str | None = None,
    sha256: str | None = None,
    currency: str = "USD",
) -> ImportJob:
    """Apply an approved plan in a single transaction.

    Either the whole workbook lands or none of it does. A half-applied price
    list is worse than a rejected one: it would leave some variants on new
    prices and some on old, with nothing to indicate which.

    Prices are never updated in place. A superseded row gets ``effective_to``
    set to the day before the new one starts, and a new row is inserted.
    """
    tiers = price_tier_map(session)
    job = ImportJob(
        file_name=file_name,
        sheet_name=sheet_name,
        storage_key=storage_key,
        sha256=sha256,
        effective_from=plan.effective_from,
        status=ImportJobStatus.PENDING,
        uploaded_by_id=getattr(user, "id", None),
        rows_total=len(plan.plans),
    )
    session.add(job)
    session.flush()

    created = updated = skipped = failed = 0

    try:
        for row_plan in plan.plans:
            record = ImportRow(
                import_job_id=job.id,
                source_row_no=row_plan.row.source_row_no,
                section_label=row_plan.row.section_label,
                raw_json={k: _jsonable(v) for k, v in row_plan.row.raw.items()},
                normalized_json=(
                    row_plan.row.parsed.model_dump(mode="json")
                    if row_plan.row.ok else None
                ),
                action=row_plan.action,
                status=row_plan.status,
                error_message=row_plan.message,
            )

            if row_plan.is_error:
                failed += 1
                session.add(record)
                continue

            if row_plan.action is ImportRowAction.SKIP:
                skipped += 1
                record.product_variant_id = row_plan.variant_id
                session.add(record)
                continue

            parsed = row_plan.row.parsed
            variant = _upsert_variant(session, parsed, row_plan)
            price_ids = _write_prices(
                session, variant, parsed, plan.effective_from, tiers, job, currency,
                only_tiers=set(row_plan.price_changes),
            )

            record.product_variant_id = variant.id
            record.created_price_ids = price_ids
            session.add(record)

            if row_plan.action is ImportRowAction.CREATE:
                created += 1
            else:
                updated += 1

        job.rows_created = created
        job.rows_updated = updated
        job.rows_skipped = skipped
        job.rows_failed = failed
        job.status = ImportJobStatus.COMMITTED
        job.finished_at = dt.datetime.now(dt.UTC)
        job.summary_json = {
            "counts": plan.counts(),
            "blocks": [
                {
                    "header_row": b.header_row,
                    "rows": b.row_count,
                    "section_label": b.section_label,
                }
                for b in plan.blocks
            ],
            "terms_found": plan.terms,
            "currency": currency,
        }

        record_audit(
            session, user, AuditAction.PRICE_LIST_IMPORTED, EntityType.IMPORT_JOB, job.id,
            new_value={
                "file": file_name,
                "sheet": sheet_name,
                "effective_from": plan.effective_from.isoformat(),
                "created": created, "updated": updated,
                "skipped": skipped, "failed": failed,
            },
        )
        session.flush()
        log.info(
            "Import committed: %s created=%d updated=%d skipped=%d failed=%d",
            file_name, created, updated, skipped, failed,
        )
        return job

    except Exception as exc:
        session.rollback()
        # Record the failure in its own transaction — the rollback above
        # discarded the job row along with everything else.
        failed_job = ImportJob(
            file_name=file_name,
            sheet_name=sheet_name,
            storage_key=storage_key,
            sha256=sha256,
            effective_from=plan.effective_from,
            status=ImportJobStatus.FAILED,
            uploaded_by_id=getattr(user, "id", None),
            rows_total=len(plan.plans),
            error_text=f"{type(exc).__name__}: {exc}",
        )
        session.add(failed_job)
        record_audit(
            session, user, AuditAction.IMPORT_FAILED, EntityType.IMPORT_JOB,
            new_value={"file": file_name}, reason=str(exc)[:500],
        )
        session.commit()
        log.exception("Import failed for %s", file_name)
        raise


def _jsonable(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (dt.date, dt.datetime)):
        return value.isoformat()
    return str(value)


def _upsert_variant(
    session: Session, parsed: PriceRowInput, row_plan: RowPlan
) -> ProductVariant:
    """Find or create the product and its board-quality variant.

    A product is the geometry; the variant carries board quality and case pack.
    Two rows with the same size and different qualities therefore share a
    product and produce two variants — which is exactly what the reference
    workbook's 14"–18" sizes require.
    """
    if row_plan.variant_id is not None:
        return session.get(ProductVariant, row_plan.variant_id)

    from modules.repositories import find_product_by_size

    product = find_product_by_size(session, parsed.product)
    if product is None:
        item_number = _item_number_for(parsed)
        product = Product(
            item_number=_unique_item_number(session, item_number),
            name=parsed.product,
            size_label=parsed.product,
            depth_in=coerce_decimal(re.sub(r"[^\d.]", "", parsed.depth or "")) or None,
            flute=parsed.flute,
            category="White Boxes",
            unit_of_measure="PACK",
        )
        session.add(product)
        session.flush()

    variant = ProductVariant(
        product_id=product.id,
        variant_item_number=_unique_variant_number(
            session, _variant_item_number(parsed, product.item_number)
        ),
        board_quality=parsed.board_quality,
        case_pack=parsed.case_pack,
    )
    session.add(variant)
    session.flush()
    return variant


def _unique_item_number(session: Session, base: str) -> str:
    from sqlalchemy import select

    candidate, suffix = base, 1
    while session.execute(
        select(Product.id).where(Product.item_number == candidate)
    ).first():
        suffix += 1
        candidate = f"{base}-{suffix}"
    return candidate


def _unique_variant_number(session: Session, base: str) -> str:
    from sqlalchemy import select

    candidate, suffix = base, 1
    while session.execute(
        select(ProductVariant.id).where(ProductVariant.variant_item_number == candidate)
    ).first():
        suffix += 1
        candidate = f"{base}-{suffix}"
    return candidate


def _write_prices(
    session: Session,
    variant: ProductVariant,
    parsed: PriceRowInput,
    effective_from: dt.date,
    tiers: dict[str, Any],
    job: ImportJob,
    currency: str,
    only_tiers: set[str] | None = None,
) -> list[int]:
    """Insert new prices, superseding any that are currently in force.

    ``only_tiers`` restricts the write to the tiers the plan found changed, so
    re-importing a workbook in which one price moved does not create redundant
    history rows for the tiers that did not. The plan is the single authority on
    what changed; this function does not re-decide it.

    Both the pack and the piece price are written as they appear in the
    workbook. The piece price is derived only when the column is absent — the
    two columns disagree by up to a rounding unit on real data and neither is
    authoritative over the other.
    """
    from modules.repositories import get_effective_price

    from modules.repositories import get_latest_price

    written: list[int] = []
    for tier_code, (pack, piece) in parsed.tier_prices().items():
        tier = tiers.get(tier_code)
        if tier is None or pack is None:
            continue
        if only_tiers is not None and tier_code not in only_tiers:
            continue

        # Deliberately the *latest* price rather than the one effective on the
        # new date: a lookup by date cannot see a price that starts later, and
        # backdating over one would leave two overlapping open-ended rows.
        latest = get_latest_price(session, variant.id, tier_code, currency)
        if latest is not None:
            if latest.effective_from >= effective_from:
                raise ValueError(
                    f"{variant.variant_item_number} already has a {tier_code} price "
                    f"effective {latest.effective_from}. Choose an effective date after "
                    f"that — rewriting the existing price is not possible."
                )
            if latest.effective_to is None or latest.effective_to >= effective_from:
                supersede_price(latest, effective_from)

        price = ProductPrice(
            product_variant_id=variant.id,
            price_tier_id=tier.id,
            price_per_pack=pack,
            price_per_piece=(
                piece if piece is not None
                else (pack / Decimal(parsed.case_pack)).quantize(Decimal("0.000001"))
            ),
            currency=currency,
            effective_from=effective_from,
            source_workbook_name=job.file_name,
            source_sheet_name=job.sheet_name,
            source_row_no=parsed.source_row_no,
            import_job_id=job.id,
            created_by_id=job.uploaded_by_id,
            is_active=True,
        )
        session.add(price)
        session.flush()
        written.append(price.id)
    return written
